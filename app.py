from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from datetime import datetime
import json
import os
from werkzeug.security import generate_password_hash, check_password_hash
import random
import csv

# Determine the base directory (where this script is located)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

template_folder = os.path.join(BASE_DIR, 'templates')
app = Flask(__name__, template_folder=template_folder)

# Fixed secret key for session persistence
app.secret_key = 'reporter-holiday-shifts-secret-key-2025'

# Data storage
DATA_DIR = os.path.join(BASE_DIR, 'data')
os.makedirs(DATA_DIR, exist_ok=True)

# Use shared reporters file from weekend_reporter if available
WEEKEND_REPORTER_DIR = os.path.join(os.path.dirname(BASE_DIR), 'weekend_reporter')
SHARED_REPORTERS_FILE = os.path.join(WEEKEND_REPORTER_DIR, 'data', 'reporters.json')

# If weekend_reporter's file exists, use it; otherwise use local
if os.path.exists(SHARED_REPORTERS_FILE):
    REPORTERS_FILE = SHARED_REPORTERS_FILE
    print(f"Using shared reporters file from weekend_reporter: {SHARED_REPORTERS_FILE}")
else:
    REPORTERS_FILE = os.path.join(DATA_DIR, 'reporters.json')
    print(f"Using local reporters file: {REPORTERS_FILE}")

SIGNUPS_FILE = os.path.join(DATA_DIR, 'signups.json')
SETTINGS_FILE = os.path.join(DATA_DIR, 'settings.json')
ASSIGNMENTS_FILE = os.path.join(DATA_DIR, 'assignments.json')
HOLIDAYS_FILE = os.path.join(DATA_DIR, 'holidays.json')

# Hard-coded deadline
DEADLINE = datetime(2025, 12, 8, 12, 0, 0)  # December 8, 2025 at noon

# Initialize data files
def init_data_files():
    # Always load reporters from CSV if it exists (to catch any CSV updates)
    reporters = {}
    csv_loaded = False
    
    # Manager account
    reporters['admin'] = {
        'name': 'Admin',
        'is_manager': True,
        'password': generate_password_hash('admin123')
    }
    
    # Load from reporter_credentials.csv (local copy or parent directory)
    csv_path = os.path.join(BASE_DIR, 'reporter_credentials.csv')
    if not os.path.exists(csv_path):
        csv_path = os.path.join(os.path.dirname(BASE_DIR), 'weekend_reporter', 'reporter_credentials.csv')
    
    if os.path.exists(csv_path):
        # Try multiple encodings
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                with open(csv_path, 'r', encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        username = row['Username']
                        reporters[username] = {
                            'name': row['Name'],
                            'is_manager': False,
                            'password': generate_password_hash(row['Password']),
                            'email': row['Email']
                        }
                csv_loaded = True
                print(f"Loaded {len(reporters)-1} reporters from CSV: {csv_path}")
                break  # Success, stop trying encodings
            except UnicodeDecodeError:
                continue  # Try next encoding
    
    # If we loaded from CSV AND reporters.json doesn't exist, save it
    # Don't overwrite existing reporters.json (preserves password changes)
    if csv_loaded and not os.path.exists(REPORTERS_FILE):
        save_json(REPORTERS_FILE, reporters)
        print(f"Created reporters.json with {len(reporters)} accounts from CSV")
    
    if not os.path.exists(SIGNUPS_FILE):
        save_json(SIGNUPS_FILE, {})
    
    if not os.path.exists(SETTINGS_FILE):
        save_json(SETTINGS_FILE, {
            'deadline': DEADLINE.isoformat(),
            'is_locked': False
        })
    
    if not os.path.exists(ASSIGNMENTS_FILE):
        save_json(ASSIGNMENTS_FILE, {})
    
    # Copy holidays.json from git if it doesn't exist in data directory
    # (This handles the case where data/ is a mounted disk that overlays git files)
    if not os.path.exists(HOLIDAYS_FILE):
        # Check if there's a holidays.json in the project root or a backup location
        backup_path = os.path.join(BASE_DIR, 'holidays_backup.json')
        if os.path.exists(backup_path):
            import shutil
            shutil.copy(backup_path, HOLIDAYS_FILE)
            print(f"Copied holidays.json from backup to {HOLIDAYS_FILE}")

# Helper functions
def load_json(filepath):
    with open(filepath, 'r') as f:
        return json.load(f)

def save_json(filepath, data):
    with open(filepath, 'w') as f:
        json.dump(data, f, indent=2)

# Initialize after defining helper functions
init_data_files()

# Template filters
@app.template_filter('format_date')
def format_date(date_str):
    """Convert YYYY-MM-DD to 'Month DD, YYYY'"""
    try:
        dt = datetime.strptime(date_str, '%Y-%m-%d')
        return dt.strftime('%B %d, %Y')
    except:
        return date_str

def get_reporters():
    return load_json(REPORTERS_FILE)

def get_signups():
    return load_json(SIGNUPS_FILE)

def get_settings():
    return load_json(SETTINGS_FILE)

def get_assignments():
    return load_json(ASSIGNMENTS_FILE)

def get_holidays():
    return load_json(HOLIDAYS_FILE)

# Routes
@app.route('/')
def index():
    if 'username' in session:
        if session.get('is_manager'):
            return redirect(url_for('manager_dashboard'))
        else:
            return redirect(url_for('reporter_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.json
        username = data.get('username')
        password = data.get('password')
        
        reporters = get_reporters()
        
        if username in reporters:
            if check_password_hash(reporters[username]['password'], password):
                session['username'] = username
                session['is_manager'] = reporters[username].get('is_manager', False)
                return jsonify({'success': True, 'is_manager': session['is_manager']})
        
        return jsonify({'success': False, 'message': 'Invalid credentials'}), 401
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/manager/dashboard')
def manager_dashboard():
    if not session.get('is_manager'):
        return redirect(url_for('login'))
    
    reporters = get_reporters()
    settings = get_settings()
    signups = get_signups()
    assignments = get_assignments()
    holidays = get_holidays()
    
    # Count reporters who have submitted signups
    submitted_count = len([r for r in signups.keys() if signups[r]])
    
    # Calculate interest per shift
    shift_interest = {}
    for holiday in holidays['shifts']:
        shift_id = holiday['id']
        interested = sum(1 for reporter_shifts in signups.values() if shift_id in reporter_shifts)
        shift_interest[shift_id] = interested
    
    return render_template('manager_dashboard.html', 
                         reporters=reporters,
                         settings=settings,
                         submitted_count=submitted_count,
                         total_reporters=len([r for r in reporters.values() if not r.get('is_manager')]),
                         assignments=assignments,
                         signups=signups,
                         holidays=holidays['shifts'],
                         shift_interest=shift_interest,
                         deadline=DEADLINE)

@app.route('/reporter/dashboard')
def reporter_dashboard():
    if 'username' not in session or session.get('is_manager'):
        return redirect(url_for('login'))
    
    settings = get_settings()
    signups = get_signups()
    assignments = get_assignments()
    holidays = get_holidays()
    username = session['username']
    
    user_signups = signups.get(username, [])
    user_assignment = assignments.get(username, None)
    
    # Check if deadline has passed
    is_locked = settings.get('is_locked', False) or datetime.now() > DEADLINE
    
    return render_template('reporter_dashboard.html',
                         username=username,
                         holidays=holidays['shifts'],
                         signups=user_signups,
                         assignment=user_assignment,
                         deadline=DEADLINE,
                         is_locked=is_locked)

@app.route('/api/signups', methods=['GET', 'POST'])
def manage_signups():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 403
    
    username = session['username']
    signups = get_signups()
    settings = get_settings()
    
    # Check if locked
    is_locked = settings.get('is_locked', False) or datetime.now() > DEADLINE
    
    if request.method == 'POST':
        if is_locked and not session.get('is_manager'):
            return jsonify({'error': 'Signups are locked'}), 403
        
        data = request.json
        user_signups = data.get('signups', [])
        
        signups[username] = user_signups
        save_json(SIGNUPS_FILE, signups)
        return jsonify({'success': True})
    
    # GET
    if session.get('is_manager'):
        return jsonify(signups)
    else:
        return jsonify({username: signups.get(username, [])})

@app.route('/api/allocate', methods=['POST'])
def allocate_shifts():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    signups = get_signups()
    holidays = get_holidays()
    assignments = {}
    shift_assignments = {shift['id']: [] for shift in holidays['shifts']}
    
    # Get all reporters who signed up
    interested_reporters = [r for r in signups.keys() if signups[r]]
    
    # Randomize order for fairness
    random.shuffle(interested_reporters)
    
    # Allocate shifts - each reporter gets max 1 shift
    for reporter in interested_reporters:
        # Skip if already assigned
        if reporter in assignments:
            continue
        
        # Get their interested shifts
        reporter_interests = signups[reporter]
        
        # Find available shifts from their interests
        available = []
        for shift_id in reporter_interests:
            shift = next((s for s in holidays['shifts'] if s['id'] == shift_id), None)
            if shift and len(shift_assignments[shift_id]) < shift['slots']:
                available.append(shift_id)
        
        # If they have available shifts, randomly pick one
        if available:
            selected_shift = random.choice(available)
            assignments[reporter] = selected_shift
            shift_assignments[selected_shift].append(reporter)
    
    # Save assignments
    save_json(ASSIGNMENTS_FILE, assignments)
    
    # Lock signups
    settings = get_settings()
    settings['is_locked'] = True
    save_json(SETTINGS_FILE, settings)
    
    return jsonify({
        'success': True,
        'assignments': assignments,
        'shift_assignments': shift_assignments
    })

@app.route('/api/reset-system', methods=['POST'])
def reset_system():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.json
    confirmation = data.get('confirmation')
    
    if confirmation != 'RESET':
        return jsonify({'error': 'Invalid confirmation'}), 400
    
    try:
        # Reset signups and assignments
        save_json(SIGNUPS_FILE, {})
        save_json(ASSIGNMENTS_FILE, {})
        
        # Unlock system
        settings = get_settings()
        settings['is_locked'] = False
        save_json(SETTINGS_FILE, settings)
        
        return jsonify({'success': True, 'message': 'System reset successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/change-password', methods=['POST'])
def change_password():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 403
    
    data = request.json
    current_password = data.get('current_password')
    new_password = data.get('new_password')
    
    if not current_password or not new_password:
        return jsonify({'error': 'Missing required fields'}), 400
    
    username = session['username']
    reporters = get_reporters()
    
    # Verify current password
    if not check_password_hash(reporters[username]['password'], current_password):
        return jsonify({'error': 'Current password is incorrect'}), 401
    
    # Update password
    reporters[username]['password'] = generate_password_hash(new_password)
    save_json(REPORTERS_FILE, reporters)
    
    return jsonify({'success': True, 'message': 'Password changed successfully'})

@app.route('/api/export-excel')
def export_excel():
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        assignments = get_assignments()
        reporters = get_reporters()
        signups = get_signups()
        holidays = get_holidays()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Holiday Shifts"
        
        # Title
        ws['A1'] = 'Reuters Reporter Holiday Shifts 2025-2026'
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = ['Holiday', 'Date', 'Time', 'Assigned Reporter(s)', 'Interest Count', 'Status', 'Capacity']
        header_row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data
        row = header_row + 1
        for shift in holidays['shifts']:
            shift_id = shift['id']
            assigned = [r for r, s_id in assignments.items() if s_id == shift_id]
            interested = sum(1 for reporter_shifts in signups.values() if shift_id in reporter_shifts)
            
            # Holiday
            ws.cell(row=row, column=1).value = shift['holiday']
            
            # Date
            ws.cell(row=row, column=2).value = shift['date']
            
            # Time
            ws.cell(row=row, column=3).value = shift['time']
            
            # Assigned reporters
            if assigned:
                reporter_names = [reporters[r]['name'] for r in assigned]
                ws.cell(row=row, column=4).value = ", ".join(reporter_names)
            else:
                ws.cell(row=row, column=4).value = "VACANT"
                ws.cell(row=row, column=4).font = Font(color="FF0000", bold=True)
            
            # Interest count
            ws.cell(row=row, column=5).value = interested
            
            # Status
            filled = len(assigned)
            total = shift['slots']
            if filled >= total:
                ws.cell(row=row, column=6).value = "FILLED"
                ws.cell(row=row, column=6).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws.cell(row=row, column=6).value = f"VACANT ({total - filled})"
                ws.cell(row=row, column=6).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Capacity
            ws.cell(row=row, column=7).value = f"{filled}/{total}"
            
            row += 1
        
        # Reporter summary section
        row += 2
        ws.cell(row=row, column=1).value = "Reporter Summary"
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        summary_headers = ['Reporter', 'Expressed Interest', 'Assigned Shift', 'Status']
        for col, header in enumerate(summary_headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
        
        row += 1
        for rep_username, rep_data in reporters.items():
            if rep_data.get('is_manager'):
                continue
            
            ws.cell(row=row, column=1).value = rep_data['name']
            
            # Expressed interest
            interested = len(signups.get(rep_username, []))
            ws.cell(row=row, column=2).value = f"Yes ({interested} shifts)" if interested > 0 else "No"
            
            # Assigned shift
            if rep_username in assignments:
                shift_id = assignments[rep_username]
                shift = next((s for s in holidays['shifts'] if s['id'] == shift_id), None)
                if shift:
                    ws.cell(row=row, column=3).value = f"{shift['holiday']} - {shift['time']}"
                    ws.cell(row=row, column=4).value = "Assigned"
                    ws.cell(row=row, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                ws.cell(row=row, column=3).value = "None"
                if interested > 0:
                    ws.cell(row=row, column=4).value = "Interested but not assigned"
                    ws.cell(row=row, column=4).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    ws.cell(row=row, column=4).value = "No interest"
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 10
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'holiday_shifts_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/upload-reporters-page')
def upload_reporters_page():
    """Upload entire reporters.json from weekend_reporter (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return redirect(url_for('login'))
    return render_template('upload_reporters.html')

@app.route('/api/upload-reporters', methods=['POST'])
def upload_reporters():
    """Upload reporters.json file directly (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        new_reporters = data.get('reporters')
        
        if not new_reporters:
            return jsonify({'error': 'No reporters data provided'}), 400
        
        # Save to reporters file
        save_json(REPORTERS_FILE, new_reporters)
        
        # Verify it saved
        reloaded = get_reporters()
        
        return jsonify({
            'success': True,
            'message': 'Reporters file uploaded successfully',
            'reporters_count': len(new_reporters),
            'verified': len(reloaded) == len(new_reporters)
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/sync-passwords-page')
def sync_passwords_page():
    """Show password sync interface (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return redirect(url_for('login'))
    return render_template('sync_passwords.html')

@app.route('/api/sync-passwords', methods=['POST'])
def sync_passwords():
    """Sync passwords from weekend_reporter's reporters.json (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        data = request.json
        weekend_reporters = data.get('reporters')
        
        if not weekend_reporters:
            return jsonify({'error': 'No reporters data provided'}), 400
        
        # Load current holiday_reporter reporters
        holiday_reporters = get_reporters()
        
        # Sync/add reporters from weekend_reporter to holiday_reporter
        synced_count = 0
        added_count = 0
        for username, weekend_data in weekend_reporters.items():
            if username == 'admin':
                continue  # Skip admin account
            
            if username in holiday_reporters:
                # Update existing reporter's password
                holiday_reporters[username]['password'] = weekend_data['password']
                synced_count += 1
            else:
                # Add new reporter
                holiday_reporters[username] = {
                    'name': weekend_data['name'],
                    'is_manager': False,
                    'password': weekend_data['password'],
                    'email': weekend_data.get('email', '')
                }
                added_count += 1
        
        # Save updated reporters
        save_json(REPORTERS_FILE, holiday_reporters)
        
        return jsonify({
            'success': True,
            'message': f'Synced {synced_count} passwords, added {added_count} new reporters',
            'synced_count': synced_count,
            'added_count': added_count
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-signups')
def download_signups():
    """Download raw signups.json file (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        signups = get_signups()
        
        # Create JSON string with pretty formatting
        json_str = json.dumps(signups, indent=2)
        
        # Create BytesIO object to send as file
        from io import BytesIO
        output = BytesIO(json_str.encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/json',
            as_attachment=True,
            download_name=f'holiday_reporter_signups_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-assignments')
def download_assignments():
    """Download raw assignments.json file (ADMIN ONLY)"""
    if not session.get('is_manager'):
        return jsonify({'error': 'Unauthorized'}), 403
    
    try:
        assignments = get_assignments()
        
        # Create JSON string with pretty formatting
        json_str = json.dumps(assignments, indent=2)
        
        # Create BytesIO object to send as file
        from io import BytesIO
        output = BytesIO(json_str.encode('utf-8'))
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/json',
            as_attachment=True,
            download_name=f'holiday_reporter_assignments_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5001)
